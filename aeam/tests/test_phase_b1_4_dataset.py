"""
aeam/tests/test_phase_b1_4_dataset.py

Phase B1.4 — Dataset & Schema Registration.

Three layers, all self-contained (no Qdrant / Redis / Postgres):

1. Schema inference (aeam.ingestion.schema_inference) — column type/role/metric
   detection from DataFrames + real CSV/XLSX bytes.
2. Dataset processor (aeam.ingestion.dataset_processor.DatasetIngestJobProcessor)
   driven through the real IngestionWorker against a temp SQLite DatabaseClient
   and a temp on-disk BlobStore.
3. Routing (aeam.ingestion.routing.RoutingJobProcessor) — dispatch by parent type.
"""

from __future__ import annotations

import io

import pandas as pd
import pytest

from aeam.ingestion.schema_inference import (
    infer_schema, read_primary_table, SchemaInferenceError,
    TYPE_INTEGER, TYPE_FLOAT, TYPE_BOOLEAN, TYPE_DATETIME, TYPE_STRING,
    ROLE_METRIC, ROLE_TIMESTAMP, ROLE_IDENTIFIER, ROLE_DIMENSION,
)
from aeam.ingestion.dataset_processor import DatasetIngestJobProcessor
from aeam.ingestion.routing import RoutingJobProcessor
from aeam.ingestion.worker import IngestionWorker
from aeam.integrations.database import DatabaseClient
from aeam.storage.blob_store import LocalDiskBlobStore
from aeam.registry.models import (
    AssetStatus, Dataset, IngestionJob, JobStatus, JobType, ParentType, Source, SourceKind, Version,
)
from aeam.registry.repositories import (
    DatasetRepository, IngestionJobRepository, SchemaRepository, SourceRepository, VersionRepository,
)


# ===========================================================================
# 1. Schema inference
# ===========================================================================

def test_infer_schema_types_roles_and_metrics():
    df = pd.DataFrame({
        "user_id": [1, 2, 3],
        "region": ["EMEA", "APAC", "AMER"],
        "revenue": [10.5, 20.0, 30.25],
        "signups": [100, 200, 300],
        "active": [True, False, True],
    })
    df["event_time"] = pd.to_datetime(["2026-01-01", "2026-01-02", "2026-01-03"])

    schema = infer_schema(df, "sales")
    by = {c["name"]: c for c in schema["columns"]}

    assert by["user_id"]["type"] == TYPE_INTEGER
    assert by["user_id"]["role"] == ROLE_IDENTIFIER and by["user_id"]["is_metric"] is False
    assert by["region"]["type"] == TYPE_STRING and by["region"]["role"] == ROLE_DIMENSION
    assert by["revenue"]["type"] == TYPE_FLOAT and by["revenue"]["role"] == ROLE_METRIC
    assert by["signups"]["role"] == ROLE_METRIC
    assert by["active"]["type"] == TYPE_BOOLEAN and by["active"]["role"] == ROLE_DIMENSION
    assert by["event_time"]["type"] == TYPE_DATETIME and by["event_time"]["role"] == ROLE_TIMESTAMP

    assert set(schema["metric_columns"]) == {"revenue", "signups"}
    assert schema["row_count"] == 3
    assert schema["object_name"] == "sales"


def test_infer_schema_identifier_heuristic_and_nullable():
    # 'id' with a null upcasts to float in pandas, but stays an identifier by name.
    # 'paid' ends in 'id' textually but must NOT be treated as an identifier.
    df = pd.DataFrame({"id": [1.0, 2.0, None], "paid": [5.0, 6.0, 7.0]})
    by = {c["name"]: c for c in infer_schema(df, "t")["columns"]}
    assert by["id"]["role"] == ROLE_IDENTIFIER and by["id"]["nullable"] is True
    assert by["paid"]["role"] == ROLE_METRIC and by["paid"]["nullable"] is False


def test_infer_schema_empty_columns_raises():
    with pytest.raises(SchemaInferenceError):
        infer_schema(pd.DataFrame(), "empty")


def test_read_primary_table_csv():
    df, detail = read_primary_table(b"a,b\n1,2\n3,4\n", "csv", "t.csv")
    assert detail["format"] == "csv" and df.shape == (2, 2)


def test_read_primary_table_excel_first_non_empty_sheet():
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "Empty"  # first sheet left with no rows
    ws = wb.create_sheet("Data")
    ws.append(["x", "y"]); ws.append([1, 2]); ws.append([3, 4])
    buf = io.BytesIO(); wb.save(buf)

    df, detail = read_primary_table(buf.getvalue(), "excel", "w.xlsx")
    assert detail["format"] == "excel"
    assert detail["sheet_name"] == "Data" and detail["sheet_count"] == 2
    assert df.shape == (2, 2)


def test_read_primary_table_header_only_raises():
    with pytest.raises(SchemaInferenceError) as exc:
        read_primary_table(b"only_header\n", "csv", "t.csv")
    assert exc.value.reason == "empty_dataset"


# ===========================================================================
# 2. Dataset processor
# ===========================================================================

@pytest.fixture()
def db(tmp_path):
    client = DatabaseClient(database_url=f"sqlite:///{(tmp_path / 'b14.db').as_posix()}")
    yield client
    client.dispose()


@pytest.fixture()
def blob_store(tmp_path):
    return LocalDiskBlobStore(tmp_path / "blobs")


def _seed_dataset_upload(db, blob_store, data: bytes, filename: str,
                         status: str = AssetStatus.PENDING):
    """Mimic the ingress API for a structured upload: blob + source + dataset + version + job."""
    ref = blob_store.put(data)
    source_id = SourceRepository(db).create(Source(name="Manual Upload", kind=SourceKind.UPLOAD))
    dataset_id = DatasetRepository(db).create(Dataset(name=filename, source_id=source_id, status=status))
    VersionRepository(db).create(Version(
        parent_type=ParentType.DATASET, parent_id=dataset_id, version=1,
        content_hash=ref.content_hash, blob_ref=ref.uri, is_active=True,
    ))
    job_id = IngestionJobRepository(db).create(IngestionJob(
        job_type=JobType.INGEST, source_id=source_id,
        parent_type=ParentType.DATASET, parent_id=dataset_id,
        status=JobStatus.QUEUED, content_hash=ref.content_hash,
    ))
    return job_id, dataset_id


def _drain(db, blob_store):
    job_repo = IngestionJobRepository(db)
    processor = DatasetIngestJobProcessor(blob_store=blob_store, db=db)
    worker = IngestionWorker(job_repo=job_repo, processor=processor, poll_interval=1)
    return worker.run_once()


def test_dataset_processor_happy_path_registers_schema(db, blob_store):
    csv = b"region,revenue,units\nEMEA,1000.5,10\nAPAC,2000.0,20\nAMER,1500.25,15\n"
    job_id, dataset_id = _seed_dataset_upload(db, blob_store, csv, "sales.csv")

    assert _drain(db, blob_store) is True

    job = IngestionJobRepository(db).get(job_id)
    ds = DatasetRepository(db).get(dataset_id)
    assert job.status == JobStatus.DONE and job.progress == 100
    assert ds.status == AssetStatus.INDEXED
    assert ds.row_count == 3
    assert set(ds.metric_columns) == {"revenue", "units"}
    assert ds.schema_id and ds.last_ingested_at

    schema = SchemaRepository(db).get(ds.schema_id)
    assert schema is not None
    assert {c["name"] for c in schema.columns} == {"region", "revenue", "units"}
    assert schema.object_name == "sales.csv"


def test_dataset_processor_excel(db, blob_store):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "KPIs"
    ws.append(["metric", "value"]); ws.append(["latency", 220]); ws.append(["errors", 3])
    buf = io.BytesIO(); wb.save(buf)

    job_id, dataset_id = _seed_dataset_upload(db, blob_store, buf.getvalue(), "kpi.xlsx")
    assert _drain(db, blob_store) is True

    ds = DatasetRepository(db).get(dataset_id)
    assert ds.status == AssetStatus.INDEXED and ds.row_count == 2
    assert "value" in ds.metric_columns


def test_dataset_processor_dedup_noop_when_indexed(db, blob_store):
    job_id, dataset_id = _seed_dataset_upload(
        db, blob_store, b"a,b\n1,2\n", "x.csv", status=AssetStatus.INDEXED,
    )
    assert _drain(db, blob_store) is True
    assert IngestionJobRepository(db).get(job_id).status == JobStatus.DONE
    # No schema created for an already-indexed dataset.
    assert DatasetRepository(db).get(dataset_id).schema_id is None


def test_dataset_processor_empty_data_fails_cleanly(db, blob_store):
    job_id, dataset_id = _seed_dataset_upload(db, blob_store, b"only_header\n", "bad.csv")
    assert _drain(db, blob_store) is True

    job = IngestionJobRepository(db).get(job_id)
    ds = DatasetRepository(db).get(dataset_id)
    assert job.status == JobStatus.FAILED
    assert ds.status == AssetStatus.ERROR


def test_dataset_processor_missing_link_fails(db, blob_store):
    ref = blob_store.put(b"a,b\n1,2\n")
    job_repo = IngestionJobRepository(db)
    job_id = job_repo.create(IngestionJob(
        job_type=JobType.INGEST, status=JobStatus.QUEUED, content_hash=ref.content_hash,
    ))  # no parent link
    assert _drain(db, blob_store) is True
    assert job_repo.get(job_id).status == JobStatus.FAILED


# ===========================================================================
# 3. Routing
# ===========================================================================

class _Recorder:
    def __init__(self):
        self.seen: list[str] = []

    def __call__(self, job, job_repo):
        self.seen.append(job.job_id)


def test_routing_dispatches_by_parent_type():
    doc_proc, ds_proc = _Recorder(), _Recorder()
    router = RoutingJobProcessor(document_processor=doc_proc, dataset_processor=ds_proc)

    djob = IngestionJob(parent_type=ParentType.DOCUMENT, parent_id="d1")
    sjob = IngestionJob(parent_type=ParentType.DATASET, parent_id="s1")
    njob = IngestionJob(parent_type=None, parent_id=None)  # legacy/no parent -> document default

    router(djob, None)
    router(sjob, None)
    router(njob, None)

    assert ds_proc.seen == [sjob.job_id]
    assert doc_proc.seen == [djob.job_id, njob.job_id]


def test_routing_rejects_none_processors():
    with pytest.raises(ValueError):
        RoutingJobProcessor(document_processor=None, dataset_processor=_Recorder())
    with pytest.raises(ValueError):
        RoutingJobProcessor(document_processor=_Recorder(), dataset_processor=None)
